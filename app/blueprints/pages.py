from __future__ import annotations

import os
from datetime import date
from io import BytesIO

from flask import (
    Blueprint,
    abort,
    current_app,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)
from sqlalchemy import func

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..models import (
    Project,
    SalesDoc,
    MaterialItem,
    SubcontractorPayment,
    OtherExpense,
)

bp_pages = Blueprint("pages", __name__)

# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------
def _num(x):
    try:
        return float(x or 0)
    except Exception:
        return 0.0


def _project_totals(p: Project):
    materials = sum(_num(m.unit_price) * _num(m.qty) for m in (p.materials or []))
    subs_pay = sum(
        (_num(s.contract_amount) - _num(s.withholding_amount))
        for s in (p.subcontractors or [])
    )
    expenses = sum(_num(e.amount) for e in (p.expenses or []))
    grand = materials + subs_pay + expenses
    return materials, subs_pay, expenses, grand


def _excel_styles(ws):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="3A3A3A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def _boq_abs_path(rel_path: str) -> str | None:
    """
    rel_path like 'uploads/boq/excel/xxxx.xlsx' (stored in DB)
    return absolute path under static or None if unsafe
    """
    if not rel_path:
        return None
    rel_path = rel_path.replace("\\", "/").lstrip("/")
    if not rel_path.startswith("uploads/boq/"):
        return None
    return os.path.join(current_app.root_path, "static", rel_path)


# ------------------------------------------------------------
# Routes
# ------------------------------------------------------------
@bp_pages.route("/")
def home():
    return redirect(url_for("pages.project_list"))


# -------------------------
# Projects
# -------------------------
@bp_pages.route("/projects")
def project_list():
    q = (request.args.get("q") or "").strip()

    query = Project.query
    if q:
        like = f"%{q}%"
        query = query.filter((Project.code.ilike(like)) | (Project.name.ilike(like)))

    projects = query.order_by(Project.updated_at.desc()).limit(200).all()
    return render_template("projects/list.html", projects=projects, q=q)


@bp_pages.route("/projects/new")
def project_new():
    return render_template("projects/form_onepage.html", project=None)


@bp_pages.route("/projects/<int:pid>/edit")
def project_edit(pid: int):
    project = Project.query.get_or_404(pid)
    return render_template("projects/form_onepage.html", project=project)


@bp_pages.route("/projects/<int:pid>")
def project_view(pid: int):
    project = Project.query.get_or_404(pid)

    materials_total, subs_total, expenses_total, grand_total = _project_totals(project)

    # ✅ ดึงเอกสาร QT ที่ผูกไว้ (ถ้ามี) เพื่อโชว์ปุ่ม BOQ + ปุ่มไปหน้า QT
    qt_doc = None
    try:
        qt_doc = getattr(project, "sales_doc", None)
    except Exception:
        qt_doc = None

    return render_template(
        "projects/view.html",
        project=project,
        qt_doc=qt_doc,
        materials_total=materials_total,
        subs_total=subs_total,
        expenses_total=expenses_total,
        grand_total=grand_total,
    )


# ------------------------------------------------------------
# ✅ Download BOQ from Project
# ------------------------------------------------------------
@bp_pages.get("/projects/<int:pid>/boq/excel")
def project_download_boq_excel(pid: int):
    project = Project.query.get_or_404(pid)

    rel = (getattr(project, "boq_excel_path", None) or "").strip()
    abs_path = _boq_abs_path(rel)
    if not abs_path or not os.path.exists(abs_path):
        abort(404)

    directory = os.path.dirname(abs_path)
    filename = os.path.basename(abs_path)
    return send_from_directory(directory, filename, as_attachment=True)


@bp_pages.get("/projects/<int:pid>/boq/pdf")
def project_download_boq_pdf(pid: int):
    project = Project.query.get_or_404(pid)

    rel = (getattr(project, "boq_pdf_path", None) or "").strip()
    abs_path = _boq_abs_path(rel)
    if not abs_path or not os.path.exists(abs_path):
        abort(404)

    directory = os.path.dirname(abs_path)
    filename = os.path.basename(abs_path)
    return send_from_directory(directory, filename, as_attachment=True)


# ------------------------------------------------------------
# Vouchers Print
# - PV: ใบสำคัญจ่าย (รวมตามรายการที่ติ๊ก)
# - RR: ใบรับรองแทนใบเสร็จรับเงิน
# - PV_WHT: ใบสำคัญจ่าย (หัก ณ ที่จ่าย) -> ใช้เฉพาะผู้รับเหมาช่วงที่ติ๊ก (S:)
# ------------------------------------------------------------
@bp_pages.route("/projects/<int:pid>/vouchers/print", methods=["POST"])
def vouchers_print(pid: int):
    project = Project.query.get_or_404(pid)

    doc_type = (request.form.get("doc_type") or "PV").upper().strip()
    if doc_type not in ("PV", "RR", "PV_WHT"):
        doc_type = "PV"

    raw_ids = request.form.getlist("ids")  # expect ["M:12","S:5","E:9"]
    selected_rows = []
    seen = set()

    for token in raw_ids:
        token = (token or "").strip()
        if not token or token in seen:
            continue
        seen.add(token)

        if ":" not in token:
            continue

        kind, sid = token.split(":", 1)
        kind = kind.strip().upper()
        try:
            iid = int(sid)
        except Exception:
            continue

        # -----------------------
        # PV_WHT: รับเฉพาะ S เท่านั้น
        # -----------------------
        if doc_type == "PV_WHT":
            if kind != "S":
                continue  # ignore ไม่ใช่ผู้รับเหมาช่วง

            item = next(
                (
                    s
                    for s in (project.subcontractors or [])
                    if int(getattr(s, "id", 0) or 0) == iid
                ),
                None,
            )
            if not item:
                continue

            contract_amount = _num(getattr(item, "contract_amount", 0))
            withholding_rate = _num(getattr(item, "withholding_rate", 0))
            withholding_amount = _num(getattr(item, "withholding_amount", 0))
            net_pay = contract_amount - withholding_amount

            selected_rows.append(
                {
                    "kind": "S_WHT",
                    "sub_id": int(getattr(item, "id", 0) or 0),
                    "vendor_name": getattr(item, "vendor_name", "") or "ผู้รับเหมา",
                    "contract_amount": contract_amount,
                    "withholding_rate": withholding_rate,
                    "withholding_amount": withholding_amount,
                    "net_pay": net_pay,
                }
            )
            continue

        # -----------------------
        # โหมดเดิม: PV / RR
        # -----------------------
        if kind == "M":
            item = next(
                (
                    m
                    for m in (project.materials or [])
                    if int(getattr(m, "id", 0) or 0) == iid
                ),
                None,
            )
            if item:
                amount = _num(getattr(item, "unit_price", 0)) * _num(
                    getattr(item, "qty", 0)
                )
                selected_rows.append(
                    {
                        "kind": "M",
                        "title": "วัสดุ",
                        "particular": f"{getattr(item, 'item_name', '') or ''} ({getattr(item, 'item_code', '') or ''})",
                        "ref_no": getattr(item, "item_code", "") or "",
                        "amount": amount,
                    }
                )

        elif kind == "S":
            item = next(
                (
                    s
                    for s in (project.subcontractors or [])
                    if int(getattr(s, "id", 0) or 0) == iid
                ),
                None,
            )
            if item:
                pay = _num(getattr(item, "contract_amount", 0)) - _num(
                    getattr(item, "withholding_amount", 0)
                )
                selected_rows.append(
                    {
                        "kind": "S",
                        "title": "ผู้รับเหมาช่วง",
                        "particular": getattr(item, "vendor_name", "") or "ผู้รับเหมา",
                        "ref_no": "",
                        "amount": pay,
                        "withholding_rate": _num(getattr(item, "withholding_rate", 0)),
                        "withholding_amount": _num(
                            getattr(item, "withholding_amount", 0)
                        ),
                    }
                )

        elif kind == "E":
            item = next(
                (
                    e
                    for e in (project.expenses or [])
                    if int(getattr(e, "id", 0) or 0) == iid
                ),
                None,
            )
            if item:
                selected_rows.append(
                    {
                        "kind": "E",
                        "title": "ค่าใช้จ่ายอื่น",
                        "particular": f"{getattr(item, 'category', '') or 'อื่นๆ'} - {getattr(item, 'title', '') or ''}",
                        "ref_no": "",
                        "amount": _num(getattr(item, "amount", 0)),
                    }
                )

    if doc_type == "PV_WHT" and not selected_rows:
        return redirect(url_for("pages.project_view", pid=project.id))

    return render_template(
        "projects/vouchers_print.html",
        project=project,
        doc_type=doc_type,
        rows=selected_rows,
        today=date.today(),
    )


# -------------------------
# Dashboard (เดิม - โครงการ)
# -------------------------
@bp_pages.route("/dashboard")
def dashboard():
    today = date.today()
    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int)

    years = list(range(today.year - 5, today.year + 2))

    q = Project.query.filter(Project.start_date.isnot(None))
    q = q.filter(func.extract("year", Project.start_date) == year)
    if month:
        q = q.filter(func.extract("month", Project.start_date) == month)

    projects = q.all()

    tot_m = tot_s = tot_e = tot_g = 0.0
    rows = []

    for p in projects:
        m, s, e, g = _project_totals(p)
        tot_m += m
        tot_s += s
        tot_e += e
        tot_g += g
        rows.append(
            {
                "id": p.id,
                "code": p.code,
                "name": p.name,
                "materials": m,
                "subs": s,
                "expenses": e,
                "total": g,
            }
        )

    rows.sort(key=lambda r: r["total"], reverse=True)
    top5 = rows[:5]

    class Totals:
        def __init__(self, m, s, e, g):
            self.materials = m
            self.subs = s
            self.expenses = e
            self.grand = g

    return render_template(
        "dashboard.html",
        year=year,
        month=month,
        years=years,
        totals=Totals(tot_m, tot_s, tot_e, tot_g),
        top5=top5,
    )


# -------------------------
# Export Excel
# -------------------------
@bp_pages.route("/projects/<int:pid>/export.xlsx")
def project_export_xlsx(pid: int):
    project = Project.query.get_or_404(pid)
    materials_total, subs_total, expenses_total, grand_total = _project_totals(project)

    wb = Workbook()
    ws = wb.active
    ws.title = "Project"

    ws.append(
        [
            "รหัสโครงการ",
            "ชื่อโครงการ",
            "สถานะ",
            "ลูกค้า",
            "สถานที่",
            "วันเริ่ม",
            "วันสิ้นสุด",
            "วันทำงาน",
            "ค่าวัสดุ",
            "ผู้รับเหมาช่วง",
            "ค่าใช้จ่ายอื่น",
            "รวมทั้งหมด",
        ]
    )
    ws.append(
        [
            project.code,
            project.name,
            project.status,
            project.customer_name or "",
            project.location or "",
            str(project.start_date or ""),
            str(project.end_date or ""),
            project.work_days or 0,
            materials_total,
            subs_total,
            expenses_total,
            grand_total,
        ]
    )

    _excel_styles(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name=f"project_{project.code}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp_pages.route("/dashboard/export.xlsx")
def dashboard_export_xlsx():
    today = date.today()
    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int)

    q = Project.query.filter(Project.start_date.isnot(None))
    q = q.filter(func.extract("year", Project.start_date) == year)
    if month:
        q = q.filter(func.extract("month", Project.start_date) == month)

    projects = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    ws.append(["รหัส", "ชื่อโครงการ", "ค่าวัสดุ", "ผู้รับเหมาช่วง", "อื่นๆ", "รวม"])

    for p in projects:
        m, s, e, g = _project_totals(p)
        ws.append([p.code, p.name, m, s, e, g])

    _excel_styles(ws)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name=f"dashboard_{year}_{month or 'all'}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ------------------------------------------------------------
# Finance dashboards (แยก รายรับ / รายจ่าย)
# ------------------------------------------------------------
@bp_pages.get("/dashboard/income")
def dashboard_income():
    """Dashboard รายรับ (นับจากใบกำกับภาษี IV ที่สถานะ APPROVED)"""
    today = date.today()
    year = request.args.get("year", type=int) or today.year

    q = (
        SalesDoc.query.filter(SalesDoc.doc_type == "IV")
        .filter(SalesDoc.status == "APPROVED")
        .filter(SalesDoc.issue_date.isnot(None))
    )

    docs = q.all()

    month_totals = {m: 0.0 for m in range(1, 13)}
    month_count = {m: 0 for m in range(1, 13)}

    total_income = 0.0
    total_vat = 0.0
    total_wht = 0.0
    total_net = 0.0

    for d in docs:
        try:
            if not d.issue_date or int(d.issue_date.year) != int(year):
                continue
            m = int(d.issue_date.month)
        except Exception:
            continue

        gross = _num(getattr(d, "gross_total", 0))
        vat = _num(getattr(d, "vat_amount", 0))
        wht = _num(getattr(d, "wht_amount", 0))
        net_pay = _num(getattr(d, "grand_total", 0))

        month_totals[m] += gross
        month_count[m] += 1

        total_income += gross
        total_vat += vat
        total_wht += wht
        total_net += net_pay

    labels = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    series = [round(month_totals[m], 2) for m in range(1, 13)]
    counts = [month_count[m] for m in range(1, 13)]
    years = list(range(today.year - 5, today.year + 2))

    return render_template(
        "dashboard_income.html",
        year=year,
        years=years,
        labels=labels,
        series=series,
        counts=counts,
        total_income=round(total_income, 2),
        total_vat=round(total_vat, 2),
        total_wht=round(total_wht, 2),
        total_net=round(total_net, 2),
    )


@bp_pages.get("/dashboard/expense")
def dashboard_expense():
    """Dashboard รายจ่าย (รวม 3 ส่วน: วัสดุ + ผู้รับเหมาช่วง + ค่าใช้จ่ายอื่นๆ)
    ใช้ created_at ของรายการเป็นตัวอ้างอิงเดือน/ปี (TimestampMixin)
    """
    today = date.today()
    year = request.args.get("year", type=int) or today.year

    mats = MaterialItem.query.all()
    subs = SubcontractorPayment.query.all()
    oth = OtherExpense.query.all()

    month_totals = {m: 0.0 for m in range(1, 13)}
    month_count = {m: 0 for m in range(1, 13)}

    total_materials = 0.0
    total_subs = 0.0
    total_other = 0.0
    total_expense = 0.0

    def _acc(dt, amount):
        nonlocal total_expense
        try:
            if not dt or int(dt.year) != int(year):
                return
            m = int(dt.month)
        except Exception:
            return
        amt = _num(amount)
        month_totals[m] += amt
        month_count[m] += 1
        total_expense += amt

    for it in mats:
        dt = getattr(it, "created_at", None)
        amt = _num(getattr(it, "unit_price", 0)) * _num(getattr(it, "qty", 0))
        if dt and int(dt.year) == int(year):
            total_materials += amt
        _acc(dt, amt)

    for it in subs:
        dt = getattr(it, "created_at", None)
        amt = _num(getattr(it, "contract_amount", 0))
        if dt and int(dt.year) == int(year):
            total_subs += amt
        _acc(dt, amt)

    for it in oth:
        dt = getattr(it, "created_at", None)
        amt = _num(getattr(it, "amount", 0))
        if dt and int(dt.year) == int(year):
            total_other += amt
        _acc(dt, amt)

    labels = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    series = [round(month_totals[m], 2) for m in range(1, 13)]
    counts = [month_count[m] for m in range(1, 13)]
    years = list(range(today.year - 5, today.year + 2))

    return render_template(
        "dashboard_expense.html",
        year=year,
        years=years,
        labels=labels,
        series=series,
        counts=counts,
        total_expense=round(total_expense, 2),
        total_materials=round(total_materials, 2),
        total_subs=round(total_subs, 2),
        total_other=round(total_other, 2),
    )


# เผื่อคนเข้า url เก่า
@bp_pages.get("/dashboard/finance")
def dashboard_finance_redirect():
    return redirect(url_for("pages.dashboard_income"))
